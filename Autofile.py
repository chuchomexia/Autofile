import os
import re
import csv
import json
import shutil
import datetime
import subprocess
import pandas as pd
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Lista para almacenar información estática
STATIC_INFO = []

# Función para convertir rutas a formato Unix
def convert_to_unix_format(path):
    return path.replace('\\', '/')

# Función para validar que la carpeta tiene el nombre correcto
def is_valid_base_folder(base_folder):
    required_folder_names = [
        "3140_C09.06_CONTRATO_PRESTACION_SERVICIOS",
        "3140_C09.08_ORDEN_COMPRA",
        "3140_C09.11_ORDEN_PRESTACION_SERVICIOS",
        "3140_C09.12_ORDEN_TRABAJO",
        "3140_C09.17_ORDEN_CONSULTORIA",
        "3140_C09.33_ORDEN_SUMINISTROS"
    ]
    
    base_folder_name = os.path.basename(base_folder)
    return base_folder_name in required_folder_names

# Función para obtener el código de calidad del archivo
def get_quality_code(file_name):
    pattern = r"F[A-Z]{2}.(\d{2,})"
    match = re.search(pattern, file_name)
    return match.group() if match else ""

# Función para cargar información estática desde un archivo CSV
def load_static_info():
    static_info = []
    with open('lists.csv', newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            static_info.append(row)
    return static_info

# Función para cargar equivalencias de nombres desde un archivo JSON
def load_name_equivalences():
    with open('equivalences_names.json', 'r') as f:
        return json.load(f)

# Función para cargar equivalencias de tipologías desde un archivo JSON
def load_typology_equivalences():
    with open('equivalences_typologies.json', 'r') as f:
        return json.load(f)

# Función para cargar equivalencias de subseries desde un archivo JSON
def load_subseries_equivalences():
    with open('equivalences_subseries.json', 'r') as f:
        return json.load(f)

# Cargar el archivo de configuración
with open('configuration.json', 'r', encoding='utf-8') as config_file:
    config_data = json.load(config_file)

# Función para generar el marco de datos del usuario
def generate_user_data_frame(user, base_folder, name_equivalences, typology_equivalences, initial_order=1):
    user_folder = os.path.join(base_folder, user)
    expedient_order = initial_order

    existing_excel_path = os.path.join(user_folder, f'{user}.xlsx')
    if os.path.exists(existing_excel_path):
        while True:
            response = input(f"Ya hay un Excel para {user}. ¿Lo cambiamos? (s/n): ")
            if response.lower() == 's':
                break
            elif response.lower() == 'n':
                print("Vale. El expediente se omitirá.")
                return None, None
            else:
                print("¡Ups! Respuesta no válida. Por favor ingresa 's' para cambiarlo o 'n' para omitirlo.")

    pdf_info, document_order, main_file_date = process_pdf(user_folder, 0, None, name_equivalences, typology_equivalences, STATIC_INFO)
    df = pd.DataFrame(pdf_info)

    df = df.apply(update_value, args=(name_equivalences, typology_equivalences), axis=1)

    expedient_closure_date = max(pdf_info, key=lambda x: x['Fecha de creación del documento'])['Fecha de creación del documento']
    total_pages = pdf_info[-1]['Página fin']

    expedient_info = {'Código Unidad': get_unit_code(base_folder),
                        'Nombre Unidad': config_data['nombre_unidad'],
                        'Código Serie': get_series_code(base_folder),
                        'Nombre Serie': config_data['nombre_serie'],
                        'Código Subserie': get_subseries_code(base_folder),
                        'Nombre Subserie': get_subseries_name(base_folder),
                        'Nombre del expediente': get_expedient_name(user),
                        'Descripción del contenido 1': get_content_description(1),
                        'Descripción del contenido 2': get_content_description(2),
                        'Descripción del contenido 3': get_content_description(3),
                        'Fecha cierre expediente': expedient_closure_date,
                        'Orden de expediente': expedient_order,
                        'Total páginas': total_pages,
                        'Objeto inventario': config_data['objeto_inventario'],
                        'Fecha inicial': get_contract_date("Ingresa la fecha de inicio del contrato (DDMMAAAA): "),
                        'Fecha final': get_contract_date("Ingresa la fecha de finalización del contrato (DDMMAAAA): "),
                        'Frecuencia consulta': config_data['frecuencia_consulta'],
                        'Soporte': config_data['soporte'],
                        'Nombre responsable entrega': config_data['nombre_responsable_entrega'],
                        'Cargo responsable Entrega': config_data['cargo_responsable_entrega'],
                        'Fecha entrega': config_data['fecha_entrega'],
                        'Nombre responsable recibido': config_data['nombre_responsable_recibido'],
                        'Cargo responsable recibido': config_data['cargo_responsable_recibido'],
                        'Fecha de recibido': config_data['fecha_recibido'],
                        'Nombre Unidad que recibe': config_data['nombre_unidad_recibe'],
                        'Tipo de Expediente': config_data['tipo_expediente'],
                        'Acceso': config_data['acceso_expediente'],
                        'Observaciones': ''}

    df_expedient = pd.DataFrame([expedient_info])
    return df, df_expedient

# Función para procesar archivos PDF
def process_pdf(user_folder, document_order, previous_main_file_date, name_equivalences, typology_equivalences, static_info):
    pdf_info = []
    start_page = 1
    previous_end_page = 0

    for pdf_file in sorted(os.listdir(user_folder)):
        if not pdf_file.endswith('.pdf'):
            continue

        full_path = os.path.join(user_folder, pdf_file)
        try:
            with open(full_path, 'rb') as file:
                pdf_reader = PdfReader(file)
                pages = len(pdf_reader.pages)

                start_page = previous_end_page + 1
                end_page = start_page + pages - 1

                file_date = None
                date_match = re.search(r"_(\d{8})_", pdf_file)
                if date_match:
                    file_date = date_match.group(1)

                if "Anexo" in pdf_file:
                    if previous_main_file_date:
                        file_date = previous_main_file_date
                    else:
                        raise ValueError(f"No se han diligenciado correctamente las fechas en el expediente {pdf_file}")
                else:
                    previous_main_file_date = file_date

                document_order += 1

                document_name = name_equivalences.get(pdf_file, '')
                typology_document = typology_equivalences.get(pdf_file, '')
                pdf_info.append({
                    'Nombre del archivo': pdf_file,
                    'Nombre del documento': document_name,
                    'Tipología documental': typology_document,
                    'Fecha de creación del documento': file_date,
                    'Fecha incorporación expediente': file_date,
                    'Orden documento expediente': document_order,
                    'Página inicio': start_page,
                    'Página fin': end_page,
                    'Origen': config_data['origen'],
                    'Acceso': config_data['acceso'],
                    'Idioma': config_data['idioma'],
                    'Autor': config_data['autor'],
                    'Código calidad': get_quality_code(pdf_file),
                    'Numero': '',
                    'Año': '',
                    'Metadato 1': '',
                    'Metadato 2': ''
                })

                previous_end_page = end_page

        except Exception as e:
            print(f"No se pudo leer el archivo {pdf_file}: {str(e)}")

    return pdf_info, document_order, previous_main_file_date

# Función para actualizar el valor de una fila
def update_value(row, name_equivalences, typology_equivalences):
    file_name = row['Nombre del archivo']
    for name, value in name_equivalences.items():
        if name in file_name:
            row['Nombre del documento'] = value
            break
    for name, typology in typology_equivalences.items():
        if name in file_name:
            row['Tipología documental'] = typology
            break
    return row

# Función para obtener el código de unidad
def get_unit_code(base_folder_name):
    pattern = r"^\d{4}"
    match = re.search(pattern, base_folder_name)
    return match.group() if match else "3140"

# Función para obtener el código de serie
def get_series_code(base_folder_name):
    pattern = r"C\d{2}"
    match = re.search(pattern, base_folder_name)
    return match.group() if match else "C09"

# Función para obtener el nombre de expediente
def get_expedient_name(user_folder_name):
    return user_folder_name

# Función para obtener el código de subserie
def get_subseries_code(base_folder_name):
    pattern = r"C\d{2}.\d{2}"
    match = re.search(pattern, base_folder_name)
    return match.group() if match else "C09.11"

# Función para obtener el nombre de la subserie
def get_subseries_name(base_folder_name):
    subseries_equivalences = load_subseries_equivalences()
    subseries_name = "Orden de Prestación de Servicios"

    pattern = r"C\d{2}.\d{2}"
    match = re.search(pattern, base_folder_name)
    if match:
        subseries_code = match.group()
        subseries_name = subseries_equivalences.get(subseries_code, "Orden de Prestación de Servicios")

    return subseries_name

# Función para obtener la descripción del contenido
def get_content_description(number):
    description = ""
    if number == 1:
        description = input("Documento con CC o NIT (CC XXXXXXXX): ")
    elif number == 2:
        description = "Nombre " + input("Nombre completo contratista: ")
    elif number == 3:
        description = "$ " + input("Valor total del contrato: ")
    return description

# Función para obtener la fecha del contrato
def get_contract_date(message):
    valid_date = False
    while not valid_date:
        date = input(message)
        try:
            date_obj = datetime.datetime.strptime(date, '%d%m%Y')
            valid_date = True
        except ValueError:
            print("¡Ups! Fecha ingresada no válida. Ingresa la fecha en formato DDMMAAAA.")
    return date_obj.strftime('%Y%m%d')

# Función para abrir un archivo PDF
def open_pdf(pdf_name):
    subprocess.Popen(['start', '', pdf_name], shell=True)

# Función para cerrar Adobe Acrobat
def close_pdf():
    with open(os.devnull, 'w') as devnull:
        subprocess.Popen(['taskkill', '/F', '/IM', 'Acrobat.exe'], stdout=devnull, stderr=devnull)

# Función para guardar en Excel
def save_to_excel(user, user_folder, result_folder, df, df_expedient):
    workbook = Workbook()
    lists_sheet = workbook.active

    for row_index, row in enumerate(STATIC_INFO, start=1):
        for col_index, value in enumerate(row, start=1):
            lists_sheet.cell(row=row_index, column=col_index, value=value)

    xlsx_output_path = os.path.join(result_folder, f'{user}.xlsx')
    with pd.ExcelWriter(xlsx_output_path, engine='openpyxl') as writer:
        workbook = writer.book

        workbook.create_sheet(title='metadatos_expediente', index=0)
        workbook.create_sheet(title='metadatos_tipologia_documental', index=1)
        workbook.create_sheet(title='Listas', index=2)

        df.to_excel(writer, sheet_name='metadatos_tipologia_documental', index=False)
        df_expedient.to_excel(writer, sheet_name='metadatos_expediente', index=False)

        worksheet_lists = workbook['Listas']
        for row_index, row_data in enumerate(STATIC_INFO, start=1):
            for col_index, value in enumerate(row_data.values(), start=1):
                worksheet_lists.cell(row=row_index, column=col_index, value=value)

        for sheet_name in writer.book.sheetnames:
            worksheet = writer.book[sheet_name]
            for column_cells in worksheet.columns:
                for cell in column_cells:
                    cell.font = None
                    cell.border = None
                    cell.alignment = Alignment(horizontal='left')
            for column in df.columns:
                max_length = 0
                column_letter = get_column_letter(df.columns.get_loc(column) + 1)
                for cell in worksheet[column_letter]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width

        # Convertir datos a formato numérico en Excel
        for col in ['A', 'J', 'K', 'O', 'P', 'U', 'X']:  # Columnas que se convertirán a formato numérico
            worksheet_expedient = writer.sheets['metadatos_expediente']
            for cell in worksheet_expedient[col]:
                if cell.value:
                    try:
                        cell.value = float(cell.value)
                    except ValueError:
                        pass
                cell.alignment = Alignment(horizontal='right')  # Mover esta línea aquí dentro del bucle

        for col in ['D', 'E']:  # Columnas que se convertirán a formato numérico
            worksheet_tipologia = writer.sheets['metadatos_tipologia_documental']
            for cell in worksheet_tipologia[col]:  
                if cell.value:
                    try:
                        cell.value = float(cell.value)
                    except ValueError:
                        pass
                cell.alignment = Alignment(horizontal='right')  # Mover esta línea aquí dentro del bucle

        # Ajustar ancho de columnas para hoja metadatos_expediente
        for column_letter in ['R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB']:
            max_length = 0
            for cell in worksheet_expedient[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet_expedient.column_dimensions[column_letter].width = adjusted_width

    print(f"Se ha generado el archivo Excel para el expediente {user} en: {xlsx_output_path}")

    input("Presiona Enter para cerrar el PDF")

    close_pdf()

# Función principal
def main():
    base_folder = input("¡Hola! Por favor ingresa la carpeta de contratos con los que trabajaremos hoy: ")
    base_folder = convert_to_unix_format(base_folder)

    if not is_valid_base_folder(base_folder):
        print("¡Ups! La carpeta que ingresaste parece ser la incorrecta :c Deberías revisarla y estar más pendiente de tu trabajo.")
        return

    result_folder = os.path.join(base_folder, 'Excel')
    if not os.path.exists(result_folder):
        os.makedirs(result_folder)

    name_equivalences = load_name_equivalences()
    typology_equivalences = load_typology_equivalences()
    global STATIC_INFO
    STATIC_INFO = load_static_info()

    initial_order_input = input("Orden del expediente: ")
    if initial_order_input.strip():
        initial_order = int(initial_order_input)
    else:
        initial_order = 1

    for user in sorted(os.listdir(base_folder)):
        user_folder = os.path.join(base_folder, user)
        if not os.path.isdir(user_folder) or user == 'Excel':
            continue

        if os.path.isdir(user_folder) and user != 'Excel':
            pdf_name = None
            for pdf_file in os.listdir(user_folder):
                if "FCO.66_Acta_" in pdf_file and pdf_file.endswith(".pdf"):
                    pdf_name = os.path.join(user_folder, pdf_file)
                    break

            if pdf_name:
                open_pdf(pdf_name)
            else:
                print("No se encontró ningún PDF para abrir.")

            # Preguntar al usuario si desea analizar el expediente
            analyze_expedient = input(f"¿Deseas analizar el expediente {initial_order}. {user}? (S/N): ")
            if analyze_expedient.lower() != 's':
                print(f"El expediente '{user}' será omitido.")
                continue  # Saltar al siguiente expediente

            # Mostrar información al usuario
            print(f"Iniciando diligenciamiento del expediente {initial_order}. {user}...")

            df, df_expedient = generate_user_data_frame(user, base_folder, name_equivalences, typology_equivalences, initial_order)

            if df is not None and df_expedient is not None:
                save_to_excel(user, user_folder, result_folder, df, df_expedient)
                copy_excel_to_user_folder(user, base_folder, result_folder)

            initial_order += 1

            expedient_marked = False

            while True:
                response = input("¿Lo revisamos luego? (S/N): ")
                if response.lower() == 's':
                    with open(os.path.join(base_folder, 'expedientes_por_revisar.txt'), 'a') as report:
                        report.write(user + '\\n')
                    expedient_marked = True
                    break
                elif response.lower() == 'n':
                    break
                else:
                    print("¡Ups! Respuesta no válida. Por favor ingresa 'S' para marcarlo o 'N' para omitir.")

            # response = input("¿Continuamos? (S/N): ")
            # if response.lower() != 's':
            #     break

# Función para copiar el archivo Excel a la carpeta del usuario
def copy_excel_to_user_folder(user, base_folder, result_folder):
    user_folder = os.path.join(base_folder, user)
    if not os.path.exists(user_folder):
        os.makedirs(user_folder)
    
    excel_file_path = os.path.join(result_folder, f'{user}.xlsx')
    destination_path = os.path.join(user_folder, f'{user}.xlsx')
    
    shutil.copy(excel_file_path, destination_path)
    print(f"Archivo Excel guardado también en la carpeta del expediente {user}")

if __name__ == "__main__":
    main()
